"""Extract the experiment data stored in .json files under a directory, and summarize them into an Excel file."""
import os
import json
import pandas as pd
import openpyxl

def get_json_paths(directory, suffix='result.json'):
    """
    Recursively find all result json files with certain suffix under a directory.

    Args:
        directory (str): directory to search for json files.
        suffix (str): suffix of the json files to search for.
    
    Returns:
        our (list of str): json file paths.
    """
    json_paths = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(suffix):
                json_paths.append(os.path.join(root, file))
    return json_paths


def extract_data_from_json(json_paths):
    """
    Extract data from a list of.json file paths.

    Args:
        json_paths (list of str): json file paths.
    
    Returns:
        return (list): List of dictionaries, where each dictionary contains the data of one experiment.
    """
    data = []
    for json_path in json_paths:
        print('Extracting data from {}'.format(json_path))
        with open(json_path, 'r', encoding='utf-8') as file:
            result_dict = json.load(file) # dict
            row = {
                'experiment_name': str(result_dict['experiment_name']),
                'comment': str(result_dict['comment']),
                'kl_divergence_final': result_dict['kl_divergence_final'][0],
                'log_likelihood_final': result_dict['log_likelihood_final'][0],
                'mmd2_rbf_final': result_dict['mmd2_rbf_final'][0],
                'wasserstein_distance_final': result_dict['wasserstein_distance_final'][0],
                'kl_divergence_final_std': result_dict['kl_divergence_final'][1],
                'log_likelihood_final_std': result_dict['log_likelihood_final'][1],
                'mmd2_rbf_final_std': result_dict['mmd2_rbf_final'][1],
                'wasserstein_distance_final_std': result_dict['wasserstein_distance_final'][1],
                'time_string': str(result_dict['time_string']),
                }
            data.append(row)
    return data


def save_to_xlsx(data, xlsx_save_path):
    """
    Save data to an Excel file.

    Args:
        data (list): List of dictionaries, where each dictionary contains the data of one experiment.
        xlsx_save_path (str): Path to the Excel file to save data to.
    """
    sheet_title = ['experiment_name',
                   'kl_divergence_final', 'log_likelihood_final', 'mmd2_rbf_final', 'wasserstein_distance_final',
                   'kl_std', 'll_std', 'mmd2_std', 'wd_std', 
                   'comment', 'time_string']
    
    # Examine whether the summary file exists. If not, create a new one.
    if not os.path.exists(xlsx_save_path):
        df=pd.DataFrame(columns=sheet_title)
        df.to_excel(xlsx_save_path, index=False, header=True, sheet_name='Sheet1')
        print("Created a summary file at {}".format(xlsx_save_path))

    workbook = openpyxl.load_workbook(xlsx_save_path)

    sheet=workbook['Sheet1']
    for data_i in data: # Loop over different experiment data
        # If the data already exists in the sheet, skip it to avoid duplicates.
        if data_i['time_string'] in [sheet.cell(row=i, column=6).value for i in range(1, sheet.max_row + 1)]:
            continue
        sheet.append({
            "A": data_i["experiment_name"],
            "B": data_i['kl_divergence_final'],
            "C": data_i['log_likelihood_final'],
            "D": data_i['mmd2_rbf_final'],
            "E": data_i['wasserstein_distance_final'],
            "F": data_i['kl_divergence_final_std'],
            "G": data_i['log_likelihood_final_std'],
            "H": data_i['mmd2_rbf_final_std'],
            "I": data_i['wasserstein_distance_final_std'],
            "J": data_i['comment'],
            "K": data_i['time_string'],
            })
    workbook.save(xlsx_save_path) # Save all changes
    print(f"Summary file saved to {xlsx_save_path}")
    return

if __name__ == '__main__':
    directory = r"E:\PythonProjects\NCSN\NCSN_custom\results_activation"
    xlsx_save_path = 'summary11112.xlsx'
    json_paths = get_json_paths(directory)
    data = extract_data_from_json(json_paths)
    save_to_xlsx(data, xlsx_save_path)
    print('Done.')
